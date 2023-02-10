
from PyQt5.QtCore import pyqtSignal, Qt
from PyQt5.QtWidgets import QWidget, QTableWidgetItem, QMessageBox, QApplication, QAbstractItemView, QFileDialog, \
    QHeaderView
from edit_ui import Ui_Form
from openpyxl import load_workbook, Workbook
from Initialization import Initi

class EditPane(QWidget,Ui_Form,Initi):
    close_flag =1 #默认是保存的  #所有文件的  #对比sheet1和ui表格数据看是否一样
    save_or_not =1  # 默认已经保存 下板部分
    edit_what_flag =0 #默认是编辑上半部分

    tips_about_all_save_signal = pyqtSignal()
    goback_menupane_signal = pyqtSignal()
    gobak_viewpane_signal = pyqtSignal(str,str)

    clear_import_signal = pyqtSignal(list,int)  #用来打开one_dalpog
    append_import_signal = pyqtSignal(list,int)
    #0表示在这个edit——pane导入

    # book_name = None  #文件是class还是edit_class
    # sheet_name = None   #哪个sheet

    #
    edit_pane_row = None
    del_lis = []

    # 进入这个见面也要把
    #离开这个界面也要把ui数据清除

    def __init__(self, parent=None, *args, **kwargs):
        super().__init__(parent, *args, **kwargs)
        self.setupUi(self)
        #
        # 自适应
        self.tableWidget.verticalHeader().setHidden(True)
        self.tableWidget.horizontalHeader().setSectionResizeMode(QHeaderView.Stretch)
        self.tableWidget.horizontalHeader().resizeSections(QHeaderView.ResizeToContents)
        self.allow_cao()
        #建立编辑文件
        #只是第一次使用程序需要
        # 第一次使用程序  建立class_and_sheet0 文件
        if self._app_data.value("first_use_GUI_editpane") == None:
            try:
                self._app_data.setValue("first_use_GUI_editpane", 1)
                self.build_edit_file()
            except Exception as e:
                QMessageBox.information(self, "error", "edit初始化错误 %s" % e )

    def clear_ui_data(self):
        self.tableWidget.clear()
        self.textEdit.clear()

    def goback_menupane(self):
        if self.have_data_or_not ==1: #表示有数据
            self.judge_all_save_or_not()
            if self.close_flag == 1:  # 保存了  连个sheet的内容于洋
                #清空现在的数据 为下次提供
                self.clear_edit_file()
                self.clear_ui_data()
                if self.goback_flag == 0:
                    self.goback_menupane_signal.emit()
                else:
                    self.gobak_viewpane_signal.emit(self.file_name,self.sheet_name)
            else: #未保存 询问是否保存
                # self.tips_about_all_save() #询问是否保存数据
                reply = QMessageBox.question(self, "question", "数据未保存，是否保存？", QMessageBox.Yes | QMessageBox.No)
                if reply == QMessageBox.Yes:
                    # 去调用保存在哪的dialog
                    self.tips_about_all_save_signal.emit()
                else: #用户选择不保存数据
                    #也要把数据清空
                    self.clear_ui_data()
                    self.clear_edit_file()
                    if self.goback_flag == 0:
                        self.goback_menupane_signal.emit()
                    else:
                        self.gobak_viewpane_signal.emit(self.file_name,self.sheet_name)

        else:  # 无数据
            self.clear_ui_data()
            if self.goback_flag == 0:
                self.goback_menupane_signal.emit()
            else:
                self.gobak_viewpane_signal.emit(self.file_name,self.sheet_name)

    def clear_edit_file(self):
        # 先把edit_filel里面的数据删除
        wb = load_workbook("edit_file.xlsx")
        st = wb["当前编辑"]
        st_row = st.max_row

        for i in range(1, st_row + 1):
            st.delete_rows(1)
        stx = wb["对比用的"]
        stx_row = stx.max_row
        for i in range(1, stx_row + 1):
            stx.delete_rows(1)

        wb.save("edit_file.xlsx")


    #如果全为空 就在保存所有的的时候删除这行
    def closeEvent(self,event):
        try:
            reply = QMessageBox.question(self, "question", "确定退出系统？", QMessageBox.Yes | QMessageBox.No)
            if reply == QMessageBox.Yes:
                #去看两个文件数据是否一样
                if self.have_data_or_not == 1: #有数据
                    #看是否保存了下半部分内容
                    if self.save_or_not == 1:
                        self.judge_all_save_or_not()
                        if self.close_flag == 1:  #退出
                            self.clear_edit_file() #清空edit_file数据
                            event.accept()
                        else :#数据未保存
                            #询问用户是否保存数据
                            reply = QMessageBox.question(self, "question", "数据未保存，是否保存？",
                                                         QMessageBox.Yes | QMessageBox.No)
                            if reply == QMessageBox.Yes:
                                # 去调用保存在哪的dialog
                                event.ignore()  #张氏不退出
                                self.tips_about_all_save_signal.emit()
                            else: #选择不保存
                                #清空edit_file数据
                                #
                                self.clear_ui_data()
                                self.clear_edit_file()
                                event.accept()  # 退出
                    else:
                        QMessageBox.information(self, "warning", "当前正在编辑一个详细内容，点击该按钮无效，请先保存至缓存或者放弃编辑!")
                else:
                    event.accept()#退出
            else:
                event.ignore()
        except Exception as e:
            print("closeEvent editpane1",e)

    def judge_all_save_or_not(self):
        try:
            self.close_flag = 1 #这是是同过检测两个文件
            wb = load_workbook("edit_file.xlsx")
            sheet1 = wb["对比用的"]

            #保存的话edit_file两个文件内容都会更新
            # print("sheet1>",sheet1.max_row)
            # print("ui", self.tableWidget.rowCount())

            # print(self.sheet1qw.max_row)
            # if self.sheet1qw.max_row - 1 != self.tableWidget.rowCount() :
            #     print("edit pane judge_all_save_or_not 该内容未保存111 行对应不上")
            #     # self.tips_about_all_save()
            #     self.close_flag = 0
            if self.hang - 1 != self.tableWidget.rowCount():
                print("edit pane judge_all_save_or_not 该内容未保存111 行对应不上")
                self.close_flag = 0

            if self.close_flag == 1:
                #表头 目前不对比 默认是不改变的 想改变的话只能从外边导入
                ui_row = self.tableWidget.rowCount()  # 获取当前ui表格共有多少hang
                ui_col = self.tableWidget.columnCount()  # 获取当前ui表格共有多少

                #这里只检测了除详细信息以外的其他内容
                #注意break只能跳出一层循环
                for i in range(ui_row):
                    for j in range(ui_col):
                        if self.tableWidget.item(i, j) == None :
                            if sheet1.cell(i+2,j+1).value != None :
                                print("edti pane judge_all_save_or_not 该内容未保存222 空白单元格数据不一样")
                                self.close_flag= 0
                                break
                            # continue  # 防止没数据报错
                        elif self.tableWidget.item(i, j).text().strip() == "":
                            if sheet1.cell(i + 2, j + 1).value != None:
                                print("edti pane judge_all_save_or_not 该内容未保存123 空白单元格数据不一样")
                                self.close_flag = 0
                                break
                        #这个是有数据的
                        else:
                            if str(sheet1.cell(i+2,j+1).value)  != str(self.tableWidget.item(i, j).text()):
                                print("excel",type(sheet1.cell(i+2,j+1).value))
                                print("excel",sheet1.cell(i+2,j+1).value)
                                print(type(self.tableWidget.item(i, j).text()))
                                print(self.tableWidget.item(i, j).text())
                                print("edit pane judge_all_save_or_no 该内容未保存333 有数据单元格不一样")
                                self.close_flag = 0
                                break
                    if self.close_flag == 0:
                        break
                    

            if self.close_flag == 1:
                sheet0 = wb["当前编辑"]
                #正常情况下两个sheet的列一样
                #但两个的列可能不一样
                #在上面已经判断了

                #注意excel表是多一行的
                #这里检测的是详细信息
                for i in range(2, sheet0.max_row + 1):
                    if sheet1.cell(i,sheet0.max_column) == None:
                        if sheet0.cell(i , sheet0.max_column).value != None:
                            # 提示
                            # QMessageBox.information(self, "warning", "该内容未被保存！444")
                            print("该内容未被保存！444")
                            self.close_flag = 0
                            break
                        continue  # 防止没数据报错
                    else:
                        if str(sheet0.cell(i , sheet0.max_column).value) != str(sheet1.cell(i ,sheet1.max_column).value):
                            # QMessageBox.information(self, "warning", "该内容未被保存！555")
                            print("edit pane judge_all_save_or_no 该内容未被保存！555")
                            self.close_flag = 0
                            break
        except Exception as e:
            print("edit paen judge_all_save_or_no ",e)


   #清空导入或者最佳导入
    def when_import_new_data(self):
        #清空下面数据
        self.tableWidget.setCurrentItem(None)
        self.textEdit.clear()
        self.save_or_not = 1  # 默认已经保存至缓存文件sheet0中
        self.edit_what_flag = 0  #默认编辑上面的
        self.close_flag = 0   #
        # self.

    def append_data_to(self,sheet_name):

        #要把一些参数改变

        #写在ui表格中和sheet0中就行
        #先写在sheet0中
        #首先要知道现在的
        #先写在sheet0里
        #先打开外部文件看列是否一样
        #其实只要把详细信息写入sheet0中就行
        ex_file = load_workbook(self.append_file)
        ex_sheet = ex_file[sheet_name]
        ex_sheet_row = ex_sheet.max_row
        ex_sheet_col = ex_sheet.max_column
        if self.tableWidget.columnCount() != ex_sheet_col -1:
            QMessageBox.information(self, "warning", "无法追加，因为列数不一样！")
        else :
            # 写入到ui里
            # x = 1
            first_ui_row = self.tableWidget.rowCount()
            self.tableWidget.setRowCount(first_ui_row + ex_sheet_row )
            # self.tableWidget.setColumnCount(ex_sheet_col - x)


            wb = load_workbook("edit_file.xlsx")
            sheet0 = wb["当前编辑"]
            sheet0_row = sheet0.max_row
            # sheet0_col = sheet0.max_column
            #给ui设置行
            # self.tableWidget.setRowCount(ex_sheet_row +1)
            for i in range(1,ex_sheet_row +1):
                for j in range(1,ex_sheet_col +1):
                    sheet0.cell(sheet0_row+i,j).value = ex_sheet.cell(i,j).value
                    #ui
                    #有可能外部文件有None
                    print(ex_sheet.cell(i, 1).value)
                    if ex_sheet.cell(i, j).value != None:
                        # continue
                        item = QTableWidgetItem(str(ex_sheet.cell(i, j).value)) # 获取单元格内容
                        self.tableWidget.setItem(first_ui_row + i - 1,j-1, item)
                    # else:
                    #     pass
                    else :
                        item = QTableWidgetItem("")
                        self.tableWidget.setItem(first_ui_row + i - 1, j - 1, item)

            wb.save("edit_file.xlsx")
            # 更新行
            self.hang = sheet0.max_row

    def giveup_all_cao(self):
        #如果点击了这个按钮就会把sheet1里面的内容重新写入到sheet0中，并且表格中的数据也会跟新
        #
        # QMessageBox.information(self, "QAQ", "")
        reply = QMessageBox.question(self, "question", "点击'yes'后，上次保存之后的编辑操作不会被保存，确定放弃编辑？", QMessageBox.Yes | QMessageBox.No)
        if reply == QMessageBox.Yes : #放弃编辑
            #把sheet1的内容写入到sheet0中
            #并更新ui数据

            #清除sheet0中的所有数据
            #

            wb = load_workbook("edit_file.xlsx")
            sheet0 = wb["当前编辑"]
            sheet0_row = sheet0.max_row
            for i in range(1,sheet0_row +1):
                sheet0.delete_rows(1)

            sheet1 = wb["对比用的"]
            for i in range(1,sheet1.max_row +1):
                for j in range(1,sheet1.max_column + 1):
                    sheet0.cell(i,j).value = sheet1.cell(i,j).value

            #保存
            wb.save("edit_file.xlsx")

            #更新到ui中
            self.write_GUI_data_to_ui_table("edit_file.xlsx","当前编辑")
            #

            #提示完成
            QMessageBox.information(self, "QAQ", "成功放弃编辑！")




    #编辑表格或者

    #左下方得保存按牛
    # save_concrete_btn
    #保存在哪
    #这个是保存分按钮
    def save_concrete_cao(self):
        try:
            #如果保存了就不要多次保存
            #注意一种是新建，一种是编辑
            #先看现在是编辑哪里
            if self.have_data_or_not == 1:#有数据
                if self.edit_what_flag == 1:
                    if self.save_or_not ==0:#未保存
                        bk = load_workbook("edit_file.xlsx")
                        sheet0 = bk["当前编辑"]
                        #这个下方保存按钮只会把数据写在sheet0中
                        #获取列
                        #编辑旧的
                        #首先要知道是哪个行
                        #判断是新建还是编辑有的
                        #当按下新建行时就把self.edit_pane_row置NOne
                        #self.edit_pane_row是从查看按钮那里来的
                        sheet0.cell(self.edit_pane_row + 2,sheet0.max_column).value = self.textEdit.toPlainText()
                        #新建立的
                        # st.cell(st.max_row + 1, st.max_column).value = self.textEdit.toPlainText()

                        bk.save("edit_file.xlsx")

                        #并且改掉文本内容
                        #save_or_not_label
                        self.save_or_not_label.setText("当前状态-已保存")

                        #并且设置为不可以编辑
                        self.textEdit.setReadOnly(True)
                        #是否保存标志置1 表示保存了
                        self.save_or_not =1
                        # self.now_edit_concrete = 1 #1表示保存了，现在不在编辑，
                        QMessageBox.information(self, "QAQ", "编辑详细内容操作已暂时保存至缓存文件中！")
                    else :#已经保存过了
                        # self.now_edit_concrete = 1  # 1表示保存了，现在不在编辑，
                        QMessageBox.information(self, "warning", "该内容已被保存，无需多次操作！")
                else:
                    QMessageBox.information(self, "warning", "当前状态为编辑表格，点击该按钮无效！")
            else:
                QMessageBox.information(self, "warning", "暂无数据，点击该按钮无效！")
        except Exception as e:
            QMessageBox.information(self, "error", "edit_pane保存至临时文件失败:%s" % e)


    def build_edit_file(self):
        wb = Workbook()  # wb是工作本
        sheet = wb.active  #
        sheet.title = '当前编辑'
        wb.create_sheet("对比用的")
        wb.save("./edit_file.xlsx")


    #这个是打开内部文件用的
    #把内部文件数据写入到sheet0和sheet1种
    def  when_to_this_edit_pane(self,str1,str2,int):  #只要不是初始化都要调用这个
        self.file_name = str1
        self.sheet_name = str2
        self.goback_flag = int
        #再清空一次，防止上次没有清空报错
        self.clear_edit_file()
        # if str1 == "class_and_sheet.xlsx":
        #     file_name = "正式文件"
        # else:
        #     file_name = "编辑历史"
        # self.table_or_concrete_label.setText("当前状态-编辑{0}的表格-{1}".format(file_name,str2))
        #已经保存
        self.save_or_not = 1  #下方按钮值设置为已经保存
        self.save_or_not_label.setText("当前状态-已保存")

        open_book = load_workbook(str1)
        sheet = open_book[str2]

        wk = load_workbook("edit_file.xlsx")
        # self.edit_fileqw = load_workbook("edit_file.xlsx")
        sheet0 = wk["当前编辑"]
        sheet1 = wk["对比用的"]
        # bt = self.edit_fileqw["当前编辑"]
        # btx = bk["对比用的"]
        # self.sheet1qw = self.edit_fileqw["对比用的"]


        #在edit_file里面写入数据
        # for i in range(1,sheet.max_row+1):
        #     for j in range(1,sheet.max_column+1):
        #         bt.cell(i,j).value = sheet.cell(i,j).value
        #         # btx.cell(i,j).value = sheet.cell(i,j).value
        #         self.sheet1qw.cell(i,j).value = sheet.cell(i,j).value
        for i in range(1, sheet.max_row + 1):
            for j in range(1, sheet.max_column + 1):
                sheet0.cell(i, j).value = sheet.cell(i, j).value
                sheet1.cell(i, j).value = sheet.cell(i, j).value

        # self.edit_fileqw.save("edit_file.xlsx")
        self.hang = sheet1.max_row
        wk.save("edit_file.xlsx")


    #这是所有都要调用的
    def allow_cao(self):  #每次进入这个界面都要引用的 设置为不可以编辑
        self.tableWidget.setEditTriggers(QAbstractItemView.NoEditTriggers)  #设置为不可编辑
        #清空textline数据
        self.textEdit.clear()
        #tablewidget设置成没有选中任何东西的状态
        self.tableWidget.setCurrentItem(None)
        #设置为只可以单选
        self.tableWidget.setSelectionMode(QAbstractItemView.SingleSelection)
        self.tableWidget.setSelectionBehavior(QAbstractItemView.SelectRows)  # 设置选中一整行
        #textline设置为不可以编辑
        # self.textEdit.setFocusPolicy(Qt.NoFocus)
        self.textEdit.setReadOnly(True)

    #这是打开内部文件
    def write_GUI_data_to_ui_table(self,file_class,sheet_name):  #仅是名字
        open_book = load_workbook(filename=file_class)  # 打开table工作本
        sheet = open_book[sheet_name]
        if sheet.cell(1, 1).value != None:  #
            self.have_data_or_not = 1 #表示有数据

            x = 1  # 后面的详细信息也不显示
            row = sheet.max_row
            col = sheet.max_column
            lis = []
            self.tableWidget.setRowCount(row - 1)
            self.tableWidget.setColumnCount(col - x)
            for i in range(1, row + 1):
                for j in range(1, col + 1 - x):  # 列数   #第i行，第j列
                    if i == 1:
                        lis.append(str(sheet.cell(i, j).value))
                        if j == col + 1 - x - 1:
                            self.tableWidget.setHorizontalHeaderLabels(lis)
                    else:
                        #有些excel单元格没有数据 就不用写入了
                        if sheet.cell(i, j).value != None:
                            item = QTableWidgetItem(str(sheet.cell(i, j).value) ) # 获取单元格内容
                            # item = QTableWidgetItem(str(sheet.cell(i, j).value))  # 获取单元格内
                            self.tableWidget.setItem(i - 2, j - 1, item)
        else:
            self.have_data_or_not = 0  # 0表示煤油数据
            T_lis = [""]
            self.tableWidget.setHorizontalHeaderLabels(T_lis)
            self.tableWidget.setColumnCount(1)
            self.tableWidget.setRowCount(1)
            item = QTableWidgetItem("%s暂无数据" % sheet_name)
            self.tableWidget.setItem(0, 0, item)
            self.tableWidget.resizeColumnToContents(0)  # 自动列宽

    def clear_import(self):
        #写在edit_file的内容都会被清空
        #先询问
        try:
            #应该在导入前问一下要不要保存
            #不用问直接导入新的
            reply = QMessageBox.question(self, "question", "确定清空原有数据并重新导入数据？", QMessageBox.Yes | QMessageBox.No)
            if reply == QMessageBox.Yes :
                #打开选项框
                #
                self.fl, _ = QFileDialog.getOpenFileName(self, "选取文件", "./", "Excel Files (*.xlsx)")
                # 先显示里面有哪些sheet
                # 把sheet名发出去
                if self.fl.strip() != "":
                    # 清空之前的
                    bk = load_workbook(self.fl)
                    st = bk.get_sheet_names()
                    # 所有的sheet可以通过 sheets1.title 发出去
                    self.clear_import_signal.emit(st,2)
        except Exception as e:
            QMessageBox.information(self, "error", "edit_pane清空再导入失败:%s" % e)

    #这是清空导入用的
    #只需要把导入的信息写入到sheet0种，先要把sheet0的详细信息写到sheet1种
    def one_sheet(self,sheet_name):
        try:
            #只需要把数据写入到sheet0中
            # 首先从当前的edit_pane知道哪个文件
            # 从mian里传来sheet名
            wb = load_workbook(self.fl)  # 外部文件
            st = wb[sheet_name]

            #这个仅仅是保存在edit_file中
            workbook = load_workbook('edit_file.xlsx')
            sheet0 = workbook["当前编辑"]
            # sheet1 = workbook["对比用的"]
            sheet0_row = sheet0.max_row
            # sheet1_row = sheet1.max_row
            #先把之前的清空
            for i in range(1,sheet0_row+1):
                sheet0.delete_rows(1)
            #     print(sheet0.max_row)
            # for i in range(1,sheet1_row+1):
            #     sheet1.delete_rows(1)
            #写入数据到sheet1中，第一个是ui当前的数据
            #第二是sheet0的详细信息
            # ui_row = self.tableWidget.rowCount()

            for i in range(1,st.max_row+1):
                for j in range(1,st.max_column+1):
                    sheet0.cell(i,j).value = st.cell(i,j).value
                    # sheet1.cell(i, j).value = st.cell(i, j).value

            #保存
            workbook.save("edit_file.xlsx")
            #更新行
            # self.hang = sheet0.max_row

            sheet0_row1 = sheet0.max_row
            sheet0_col1 = sheet0.max_column

            #更新数据
            #先要把之前ui的清除
            self.clear_ui_data()
            lis = []
            x =1
            if sheet0.cell(1, 1).value != None:  # 说明有数据  #注意是外部文件
                self.tableWidget.setRowCount(sheet0_row1- 1)
                self.tableWidget.setColumnCount(sheet0_col1 - x)  # 最后一个详细信息不展示
                for i in range(1, sheet0_row1 + 1):  # 把对应表格的数据拿出来
                    for j in range(1, sheet0_col1 + 1 - x):  # 列数   #第i行，第j列  #最后的详细信息不展示
                        if i == 1:
                            lis.append(str(sheet0.cell(i, j).value))
                            if j == sheet0_col1 + 1 - x - 1:  # 行都有了
                                self.tableWidget.setHorizontalHeaderLabels(lis)
                        else:
                            if sheet0.cell(i,j).value == None:
                                continue
                            item = QTableWidgetItem(str(sheet0.cell(i, j).value))  # 获取单元格内容
                            self.tableWidget.setItem(i - 2, j - 1, item)
            else:  # 空的
                T_lis=[""]
                self.tableWidget.setHorizontalHeaderLabels(T_lis)
                self.tableWidget.setColumnCount(1)
                self.tableWidget.setRowCount(1)
                item = QTableWidgetItem("%s暂无数据" % sheet_name)
                self.tableWidget.setItem(0, 0, item)
                self.tableWidget.resizeColumnToContents(0)  # 自动列宽
            QMessageBox.information(self, "QAQ", "导入成功！")
        except Exception as e:
               print("edit_pane one_sheet ",e)


    def del_row(self,del_rows):
        bk = load_workbook("edit_file.xlsx")
        sheet0 = bk["当前编辑"]
        sheet0.delete_rows(del_rows + 2)
        print("excel行",sheet0.max_row)
        bk.save("edit_file.xlsx")
        # self.sheet1q
        # self.sheet0qw = self.edit_fileqw["当前编辑"]
        # self.sheet0qw.delete_rows(del_rows + 2)
        # self.edit_fileqw.save("edit_file.xlsx")
        # print("edit del_row",self.sheet0qw.max_row)

        # self.sheetqw.max_row =
        #
        # # wb = load_workbook("edit_file.xlsx")
        # # sheet1 = wb["对比用的"]
        #
        # # 保存的话edit_file两个文件内容都会更新
        # # print("sheet1>",sheet1.max_row)
        # print("ui", self.tableWidget.rowCount())
        #
        # print(self.sheet1qw.max_row)

    def append_import(self):
        #首先要看有没有数据
        try:
            if self.have_data_or_not == 1:
                if self.save_or_not == 1:  # 编辑好了下面的
                    reply = QMessageBox.question(self, "question", "确定在原有数据上追加数据？", QMessageBox.Yes | QMessageBox.No)
                    if reply == QMessageBox.Yes:
                        # 打开选项框
                        #
                        self.append_file, _ = QFileDialog.getOpenFileName(self, "选取文件", "./", "Excel Files (*.xlsx)")
                        # 先显示里面有哪些sheet
                        # 把sheet名发出去
                        if self.append_file.strip() != "":
                            # 清空之前的
                            bk = load_workbook( self.append_file)
                            st = bk.get_sheet_names()
                            # 所有的sheet可以通过 sheets1.title 发出去
                            self.append_import_signal.emit(st, 3)  #2表示
                else:
                    QMessageBox.information(self, "warning", "当前正在编辑一个详细内容，点击该按钮无效，请先保存至缓存或放弃编辑")
            else:
                QMessageBox.information(self, "warning", "暂无数据，点击该按钮无效！")
        except Exception as e:
            print("edit pane append_import",e)

    #concrete_btn
    def edit_concrete_allow_forbid_edit(self):  #这个是设置具体的
        try:
            if self.have_data_or_not == 1:
                if self.edit_what_flag == 1:
                    # if self.save_or_not == 1: #左下方的保存了
                        #只要点击了编辑按钮就是未保存
                        #这里改一下
                        #在编辑下面的按钮如果点击上面的转化
                        # self.save_or_not =0 #未保存标志
                        #提示变成了
                        if self.textEdit.isReadOnly() == True:  #为只读
                            #可以写入
                            self.textEdit.setReadOnly(False)
                            QMessageBox.information(self, "QAQ", "详细信息可编辑！")
                            self.save_or_not_label.setText("当前状态-未保存")
                            self.save_or_not =0
                            #把它设置为未保存
                            #并且写入未保存标志中
                            #但有些人会在保存之后点击仍然去点击编辑禁止编辑按牛
                            #
                        else:
                            self.textEdit.setReadOnly(True)
                            QMessageBox.information(self, "QAQ", "详细信息不可编辑！")
                    # else :
                    #     QMessageBox.information(self, "QAQ", "请先保存或放弃刚才书写的详细信息！")
                else:
                    QMessageBox.information(self, "warning", "当前状态为编辑表格，点击该按钮无效！")
            else:
                QMessageBox.information(self, "warning", "暂无数据，点击该按钮无效！")
        except Exception as e:
            QMessageBox.information(self, "error", "edit_pane错误为%s" % e)

        #当选中某个时，下面显示详情

    def add_row_cao(self):
        try:
            if self.have_data_or_not == 1:
                if self.edit_what_flag == 0: #可以编辑表格
                    if self.tableWidget.editTriggers() ==  QAbstractItemView.CurrentChanged : #可以编辑
                        self.tableWidget.setRowCount(self.tableWidget.rowCount() + 1)
                        self.tableWidget.selectRow(self.tableWidget.rowCount() - 1)  # 注意最大一行是要减一1的
                    else:
                        QMessageBox.information(self, "warning", "表格当前状态为不可编辑！")
                else:
                    QMessageBox.information(self, "warning", "当前状态为编辑详细内容，点击该按钮无效！")
            else:
                QMessageBox.information(self, "warning", "暂无数据，点击该按钮无效！")
        except Exception as e:
            QMessageBox.information(self, "error", "edit_pane 增加行时出错，错误为:%s" % e)

    def delete_row_cao(self):
        try:
            if self.have_data_or_not == 1:
                if self.edit_what_flag == 0:  # 可以编辑表格
                    if self.tableWidget.editTriggers() == QAbstractItemView.CurrentChanged:  # 可以编辑
                        s_items = self.tableWidget.selectedItems()  # 获取当前所有选择的items
                        if s_items:
                            #
                            # self.tableWidget.setStyleSheet("selection-background-color:rgb(255,209,128)")
                            selected_rows = []  # 求出所选择的行数
                            for i in s_items:
                                row = i.row()
                                if row not in selected_rows:
                                    selected_rows.append(row)
                            for r in range(len(sorted(selected_rows))):
                                # 获取编号
                                    # 注意从0开始
                                    # print("行",self.tableWidget.currentRow())
                                    # self.identifier = self.tableWidget.item(self.tableWidget.currentRow(),self.identifier_position -1).text()
                                    # print(self.identifier)
                                self.tableWidget.removeRow(selected_rows[r] - r)  #
                                # 去往copy的文件 把对应行删除
                                # self.del_row(self.tableWidget.currentRow())
                                # print("ui行",self.tableWidget.currentRow())
                                self.del_row(selected_rows[r] - r)
                                #提示删除成功
                                QMessageBox.information(self, "QAQ", "删除成功！")

                                    # 如果有删除就建立新文件
                                    # 先把文件copy一份
                                    # 操作的是copy中的数据
                        else:  # 未选中对像
                            QMessageBox.information(self, "warning", "请先选中一个对象！")
                    else:
                        QMessageBox.information(self, "warning", "表格当前状态为不可编辑！")
                else:
                    QMessageBox.information(self, "warning", "当前状态为编辑详细内容，点击该按钮无效！")
            else:
                QMessageBox.information(self, "warning", "暂无数据，点击该按钮无效！")
        except Exception as e:
            QMessageBox.information(self, "error", "edit_pane 删除行时出错，错误为:%s" % e)

    #0表示编辑表格 edit_what_flag
    #1比啊【
    #标签table_or_concrete_label
    # 按钮edit_table_or_concrete_btn
    def edit_table_or_concrete(self):
        try:
            #先看之前有没有编辑过详细内容
            #编辑过保存了就直接改变
            #编辑过但未保存
            #未编辑未保存
            # if self.edit_concrete_or_not_flag == 1: #编辑过详细内容
                #是否保存

            # self.save_or_not = 1  # 保存标志置1
            # self.now_edit_concrete = 1 # 当前不在编辑
            if self.have_data_or_not == 1:
                if self.edit_what_flag == 1:
                    if  self.save_or_not == 0:  #未保存
                        reply = QMessageBox.question(self, "question", "刚才编辑的详细内容未保存,是否保存至对应位置？", QMessageBox.Yes | QMessageBox.No)
                        if reply ==  QMessageBox.Yes :
                            #把数据写入到edit_file中
                            # self.save_concrete_cao()
                            bk = load_workbook("edit_file.xlsx")
                            sheet0 = bk["当前编辑"]
                            # 这个下方保存按钮只会把数据写在sheet0中
                            # 获取列
                            # 编辑旧的
                            # 首先要知道是哪个行
                            # 判断是新建还是编辑有的
                            # 当按下新建行时就把self.edit_pane_row置NOne
                            # self.edit_pane_row是从查看按钮那里来的
                            sheet0.cell(self.edit_pane_row + 2, sheet0.max_column).value = self.textEdit.toPlainText()

                            self.save_or_not = 1
                            #提示刚才的详细信息保存成功
                            QMessageBox.information(self, "QAQ", "编辑详细内容操作已暂时保存至缓存文件中！")
                            self.textEdit.clear()
                            # 并且设置为不可以编辑
                            self.textEdit.setReadOnly(True)

                            self.edit_what_flag = 0
                            self.table_or_concrete_label.setText("当前状态-编辑表格")
                            self.con_lebel.setText("详细信息")
                            QMessageBox.information(self, "QAQ", "当前状态为：编辑表格！")
                            self.textEdit.clear()
                            # 并且设置为不可以编辑
                            self.textEdit.setReadOnly(True)

                            # else:
                            #这样设置，只要是点击了编辑按钮就是未保存
                                #在没有按下编辑时，显示的是已经保存
                                #按下可以编辑后，显示的是未保存
                                #此时查看按钮仍然可以点击
                                    # 有按下就清空
                    else : #已经保存
                        self.edit_what_flag = 0
                        self.table_or_concrete_label.setText("当前状态-编辑表格")
                        QMessageBox.information(self, "QAQ", "当前状态为：编辑表格！")
                        self.textEdit.clear()
                        # 并且设置为不可以编辑
                        self.textEdit.setReadOnly(True)

                # 一开始
                # if self.edit_what_flag == 0:
                else:
                    self.table_or_concrete_label.setText("当前状态-编辑详细信息")
                    self.edit_what_flag = 1
                    # 把表格设置为不可编辑状态，并且提示的也变一下
                    self.table_allow_forbid_label.setText("表格-当前不可编辑")
                    self.tableWidget.setEditTriggers(QAbstractItemView.NoEditTriggers)  # 设置为不可编辑
                    # 关于表格的任何东西点击都会跳出提示框
                    QMessageBox.information(self, "QAQ", "当前状态为：编辑详细信息！")
                    #
                # else:
                #     self.edit_what_flag = 0
                #     self.table_or_concrete_label.setText("当前状态-编辑表格")
                #     QMessageBox.information(self, "QAQ", "当前状态为：编辑表格！")
            else:
                QMessageBox.information(self, "warning", "暂无数据，点击该按钮无效！")
        except Exception as e:
            QMessageBox.information(self, "error", "edit_pane转化状态失败:%s" % e)

    #中间部分
    #标签 table_allow_forbid_label
    #按钮 table_allow_forbid_btn
    def  table_allow_forbid_cao(self):
        try:
            if self.have_data_or_not ==1:#有数据
                if self.edit_what_flag ==0: #编辑表格
                    if self.tableWidget.editTriggers() ==  QAbstractItemView.CurrentChanged :#可以编辑
                        self.tableWidget.setEditTriggers(QAbstractItemView.NoEditTriggers)
                        self.table_allow_forbid_label.setText("表格-当前不可编辑")
                        QMessageBox.information(self, "QAQ", "表格不可编辑！")
                    else :
                        self.tableWidget.setEditTriggers(QAbstractItemView.CurrentChanged)
                        self.table_allow_forbid_label.setText("表格-当前可编辑")
                        QMessageBox.information(self, "QAQ", "表格可编辑！")
                else :
                    QMessageBox.information(self, "warning", "当前状态为编辑详细信息，点击该按钮无效！")
            else :
                QMessageBox.information(self, "warning", "暂无数据，点击该按钮无效！")
        except Exception as e:
            QMessageBox.information(self, "error", "edit_pane表格允许/禁止编辑转化失败:%s" % e)

    #打开edit——file
    def to_textEdit(self):  #edit_pane_row是选中的行数
        try:
            #打开后显示已经保存
            #一旦修改就是没保
            # text = self.edit.toPlainText()
            # self.label.setText(text)
            if self.have_data_or_not  == 1:
                if self.edit_what_flag == 1 :#编辑详细内容  #并且目前不在编辑详内容
                    if self.save_or_not == 1 :#编辑好了
                        #还要近一半判断，如果现在正在编辑某格=个内容 这个不可以点击
                        #就是说选中了某行
                        row = self.tableWidget.selectedItems()
                        if len(row) > 0:  # 至少选中了一个数据
                            self.edit_pane_row = self.tableWidget.selectedItems()[0].row()
                            #要把之前的texEditt里面的内容删除
                            self.textEdit.clear()
                            # open_book = load_workbook(self.book_name)  # 哪个文件夹
                            # sheetx = open_book[self.sheet_name]
                            open_book = load_workbook("edit_file.xlsx")
                            sheetx = open_book.active
                            col = sheetx.max_column
                            if sheetx.cell(self.edit_pane_row + 2, col).value != None:
                                self.textEdit.setText(str(sheetx.cell(self.edit_pane_row + 2, col).value))
                            else :
                                #提示没有详细信息
                                QMessageBox.information(self, "warning", "暂无详细信息！")
                            x = 3 #表示编号所在的列
                            xiangxi = str(sheetx.cell(self.edit_pane_row + 2, x).value)
                            self.con_label.setText("详细信息-%s" % xiangxi)
                            self.first_text = str(sheetx.cell(self.edit_pane_row + 2, col).value)
                            self.save_or_not_label.setText("当前状态-已保存")
                        #
                        else:
                            QMessageBox.information(self, "warning", "请选中一个内容!")
                    else:
                        QMessageBox.information(self, "warning", "当前正在编辑一个详细内容，点击该按钮无效，请先保存至缓存或者放弃编辑!")
                else:
                    QMessageBox.information(self, "warning", "当前状态为编辑详表格，点击该按钮无效！")
            else:
                QMessageBox.information(self, "warning", "暂无数据，点击该按钮无效！")
        except Exception as e:
            QMessageBox.information(self, "error", "edit_pane查看详细内容时发生错误,错误为%s" %  e)

    #把数据写入到正式文件或历史种
    #dialog path 要用到
    def write_data(self,file_name):
        # edit_wb = load_workbook("edit_file.xlsx")
        # edit_sheet0 = edit_wb["当前编辑"]

        bk = load_workbook(file_name)
        bt = bk.worksheets[0]
        wb = load_workbook("edit_file.xlsx")
        sheet1 = wb["对比用的"]
        self.sheet_or_not_flag = 0  # 不存在标志
        for i in range(2, bt.max_row + 1):
            if bt.cell(i, 1).value == self.sheet_name:
                self.sheet_or_not_flag = 1   #存在这个文件
                # 把之前文件内容清空
                #
                if file_name == "class_and_sheet.xlsx" :
                    x= 3
                    if sheet1.max_row >=2 and sheet1.max_column >= x:
                        bt.cell(i,2).value = sheet1.cell(2,x).value
                    # if edit_sheet0.max_row >=2 and edit_sheet0.max_column >= x:
                    #     bt.cell(i,2).value = edit_sheet0.cell(2,x).value
                sheetx = bk[self.sheet_name]
                sheetx_row = sheetx.max_row
                for m in range(1, sheetx_row + 1):
                    sheetx.delete_rows(1)
                break

            # 这个旨在hisop
        if self.sheet_or_not_flag == 0:  # 新建
            bt.cell(bt.max_row + 1, 1).value = self.sheet_name
            bk.create_sheet(self.sheet_name)


        # for i in range(1, edit_sheet0.max_row + 1):
        #     for j in range(1, edit_sheet0.max_column + 1):
        #         bk[self.sheet_name].cell(i, j).value = edit_sheet0.cell(i, j).value
        # print(self.sheet0qw.max_row)
        # print(self.sheet0qw.max_column)
        for i in range(1, sheet1.max_row + 1):
            for j in range(1, sheet1.max_column + 1):
                bk[self.sheet_name].cell(i, j).value = sheet1.cell(i, j).value
                # bk[self.sheet_name].cell(i, j).value = self.sheet0qw.cell(i, j).value
        bk.save(file_name)


    def save_about_path(self,save_to_where_flag):
        try:
            #先把数据写在sheet0和sheet1中
            self.save_edit_btn.setFocusPolicy(Qt.StrongFocus)
            self.tableWidget.viewport().update()  # 更新数据

            # 先保存到sheet0中，
            # edit_file = load_workbook("edit_file.xlsx")
            wb = load_workbook("edit_file.xlsx")
            sheet0 = wb["当前编辑"]
            # sheet0 = edit_file["当前编辑"]
            # sheet0 = self.edit_fileqw["当前编辑"]
            # 检查ui表格是否有全空白的
            # 找到哪个空白行 然后录入数据到excel时跳过那行

            # 一般情况下ui表格和sheet0列数一样
            ui_row = self.tableWidget.rowCount()  # 获取当前ui表格共有多少hang
            ui_col = self.tableWidget.columnCount()  # 获取当前ui表格共有多少

            # 先要把sheet0之前的清空，除了详细信息
            # sheet0_row = sheet0.max_row
            # sheet0_col = sheet0.max_column

            # 清除sheet0除详细信息以外得信息
            # for i in range(1, sheet0_row + 1):  # 注意详细信息不要请了
            #     for j in range(1, sheet0_col):
            #         sheet0.cell(i, j).value = None
            # # 写入除详细信息以外的信息
            # for j in range(ui_col):
            #     sheet0.cell(1, j + 1).value = self.tableWidget.horizontalHeaderItem(j).text()
            for i in range(1, sheet0.max_row + 1):  # 注意详细信息不要请了
                for j in range(1, sheet0.max_row):
                    sheet0.cell(i, j).value = None
            # 写入除详细信息以外的信息
            for j in range(ui_col):
                sheet0.cell(1, j + 1).value = self.tableWidget.horizontalHeaderItem(j).text()
            for i in range(ui_row):
                for j in range(ui_col):
                    if self.tableWidget.item(i, j) == None:
                        # continue  # 防止没数据报错 没数据时直接跳过
                        sheet0.cell(i + 2, j + 1).value = None
                    elif self.tableWidget.item(i, j).text().strip() == "":
                        sheet0.cell(i + 2, j + 1).value = None
                    else:
                        sheet0.cell(i + 2, j + 1).value = self.tableWidget.item(i, j).text()
            # for i in range(ui_row):
            #     for j in range(ui_col):
            #         if self.tableWidget.item(i, j) == None:
            #             # continue  # 防止没数据报错 没数据时直接跳过
            #             sheet0.cell(i + 2, j + 1).value = None
            #         elif self.tableWidget.item(i, j).text().strip() == "":
            #             sheet0.cell(i + 2, j + 1).value = None
            #         else:
            #             sheet0.cell(i + 2, j + 1).value = self.tableWidget.item(i, j).text()



            # sheet1 = edit_file["对比用的"]
            sheet1 = wb["对比用的"]
            sheet1_row = sheet1.max_row
            # sheet1_row = self.sheet1qw.max_row
            # sheet1_row = edit_file["对比用的"]
            # sheet1_col = sheet1.max_column
            # for i in range(1, sheet1_row + 1):
            #     sheet1.delete_rows(1)
            for i in range(1, sheet1_row + 1):
                sheet1.delete_rows(1)
                # self.sheet1qw.delete_rows(1)
                # sheet1.delete_rows(1)
            #把sheet1里面的内容完全清空

            # for i in range(1, sheet1_row + 1):
            #     for j in range(1, sheet1_col + 1):
            #         sheet1.cell(i, j).value = None
            # for i in range(1, sheet0.max_row + 1):  # sheet0.max_row 注意这个是新的 不是之前的哪个
            #     for j in range(1, sheet0_col + 1):
            #         self.sheet1qw.cell(i, j).value = sheet0.cell(i, j).value
            # print("editpane save_about_path sheet0的行数",self.sheet0qw.max_row)
            for i in range(1, sheet0.max_row + 1):  # sheet0.max_row 注意这个是新的 不是之前的哪个
                for j in range(1, sheet0.max_column + 1):
                    sheet1.cell(i, j).value = sheet0.cell(i, j).value
                    # sheet1.cell(i, j).value = self.sheet0qw.cell(i, j).value
                    # self.sheet1qw.cell(i, j).value = self.sheet0qw.cell(i, j).value
                    # sheet1.cell(i, j).value = eet0.cell(i, j).value

            # edit_file.save("edit_file.xlsx")
            # self.edit_fileqw.save("edit_file.xlsx")
            wb.save("edit_file.xlsx")
            #获取此时的写进去的行数 防止行数报错
            self.hang = sheet1.max_row
            # 写入到正式文件或xx文件中
            #如果等于0

            #把数据写入到正式或历史文件中
            if save_to_where_flag == 0:
                self.write_data("class_and_sheet.xlsx")
            elif save_to_where_flag == 1:
                self.write_data("edit_class_and_sheet.xlsx")
            else:
                self.write_data("class_and_sheet.xlsx")
                self.write_data("edit_class_and_sheet.xlsx")


            QMessageBox.information(self, "QAQ", "保存成功!")
        except Exception as e:
            print("edit_pane save_about_path",e)

    #编辑界面
    def save_to_which_file(self):
        #这个保存是对所有编辑ide保存
        #放弃编辑 所有的编辑操作都不会被保存sheet2
        #sheet0是用来保存下面的那部分的，保存为历史或则正式就会在sheet0和sheet1写入数据 当前表格数据和sheet0的详细信息
        #看是否保存是用sheet1和当前的表格数据对比，sheet0占了一个详细信息
        #在有数据的情况下 先保存在sheet0和sheet1，和对应的文件
        #根据按钮上的文字来
        #只有退出这个界面时才有真正改变正式文件的值 因为怕中途放弃编辑，当然也可以，，，
        #现在只能咋

        #保存 把表格数据保存在sheet1中，然后直接把sheet0复制3份sheet1，sheet2，还有一个是正式文件或历史我呢见
        try:
            # self.which_file = None  # 是保存到哪个文件中
            if self.sender().text() == "保存为正式文件":
                self.which_file = "class_and_sheet.xlsx"
                self.asr = "正式文件"
            else:
                self.which_file = "edit_class_and_sheet.xlsx"
                self.asr = "编辑记录"
            reply = QMessageBox.question(self, "question", "确定要把数据保存到%s中" % self.asr, QMessageBox.Yes | QMessageBox.No)
            if reply == QMessageBox.Yes:
                if self.have_data_or_not == 1: #有数据
                    # try:
                    if self.save_or_not == 1: #保存了 #这个是下main的哪个按钮
                        self.save_edit_btn.setFocusPolicy(Qt.StrongFocus)
                        self.tableWidget.viewport().update()  # 更新数据

                        #先保存到sheet0
                        # self.edit_fileqw = load_workbook("edit_file.xlsx")
                        wb = load_workbook("edit_file.xlsx")
                        sheet0 = wb["当前编辑"]
                        # sheet0 = self.edit_fileqw["当前编辑"]

                        #检查ui表格是否有全空白的
                        #找到哪个空白行 然后录入数据到excel时跳过那行

                        #一般情况下ui表格和sheet0列数一样
                        ui_row = self.tableWidget.rowCount()  # 获取当前ui表格共有多少hang
                        ui_col = self.tableWidget.columnCount()  # 获取当前ui表格共有多少


                        #先要把sheet0之前的清空，除了详细信息
                        sheet0_row = sheet0.max_row
                        sheet0_col = sheet0.max_column


                        #有可能ui的数据有空的
                        #清除sheet0除详细信息以外得信息
                        for i in range(1,sheet0_row+1): #注意详细信息不要请了
                            for j in range(1,sheet0_col):
                                sheet0.cell(i,j).value = None
                        #写入除详细信息以外的信息
                        for j in range(ui_col):
                            sheet0.cell(1, j + 1).value = self.tableWidget.horizontalHeaderItem(j).text()
                        for i in range(ui_row):
                            for j in range(ui_col):
                                if self.tableWidget.item(i, j) == None  : #
                                    sheet0.cell(i+2,j+1).value = None
                                    # continue  # 防止没数据报错 没数据时直接跳过
                                elif self.tableWidget.item(i, j).text().strip() == "":
                                    sheet0.cell(i + 2, j + 1).value = None
                                else: #数据不为空白
                                    sheet0.cell(i+2, j + 1).value = self.tableWidget.item(i, j).text()
                                #这里错了 当ui没有数据时这里直接就不写了 这会导致缺少一行
                                #这里应该当ui为空时 写道excel用None

                        # print("ui",self.tableWidget.rowCount())
                        # print("sheet0",sheet0.max_row)

                        #主要是最后的空白行

                        #注意到sheet0的行会被清除
                        #把sheet0复制到sheet1中
                        #先要把sheet1里面的所有内容清除掉
                        # self.sheet1qw = self.edit_fileqw["对比用的"]
                        # sheet1_row = self.sheet1qw.max_row
                        # sheet1_col = self.sheet1qw.max_column
                        sheet1 = wb["对比用的"]
                        sheet1_row = sheet1.max_row

                        #把sheet1里面的数据删除
                        for i in range(1,sheet1_row+1):
                            sheet1.delete_rows(1)
                            # self.sheet1qw.delete_rows(1)

                        # for i in range(1,sheet1_row+1):
                        #     for j in range(1,sheet1_col+1):
                        #         self.sheet1qw.cell(i,j).value = None
                        for i in range(1,sheet0.max_row+1):  #sheet0.max_row 注意这个是新的 不是之前的哪个
                            for j in range(1,sheet0_col+1):
                                sheet1.cell(i, j).value = sheet0.cell(i, j).value
                                # self.sheet1qw.cell(i,j).value = sheet0.cell(i,j).value

                        # self.edit_fileqw.save("edit_file.xlsx")
                        wb.save("edit_file.xlsx")
                        self.hang = sheet1.max_row
                        #把这个行给self。行 防止检查sheet和ui 行数不一样时报错

                        # print("sheet01@",sheet0.max_row)
                        # print("sheet1",self.sheet1qw.max_row)
                        #
                        # sf = load_workbook("edit_file.xlsx")
                        # st1 = sf["当前编辑"]
                        # print("st1",st1.max_row)


                        #写入到正式文件或xx文件中
                        bk = load_workbook(self.which_file)
                        bt = bk.worksheets[0]
                        self.sheet_or_not_flag = 0  # 不存在标志
                        for i in range(2, bt.max_row + 1):
                            if bt.cell(i, 1).value == self.sheet_name:
                                self.sheet_or_not_flag = 1
                                #存在需要把现在的编号写进去
                                #只有写入到正式文件中才要
                                if self.which_file == "class_and_sheet.xlsx" :
                                    x = 3  #表示编号在第三列
                                    if sheet0.max_row >=2 and sheet0.max_column >= x:
                                        bt.cell(i,2).value = sheet0.cell(2,x).value
                                # 把之前文件内容清空
                                sheetx = bk[self.sheet_name]
                                sheetx_row = sheetx.max_row
                                # sheetx_col = sheetx.max_column
                                for i in range(1,sheetx_row +1):
                                    sheetx.delete_rows(1)
                                # for m in range(1, sheetx_row + 1):
                                #     for n in range(1, sheetx_col + 1):
                                #         sheetx.cell(m, n).value = None
                                break

                            # 这个旨在hisop
                        if self.sheet_or_not_flag == 0:  # 新建
                            bt.cell(bt.max_row + 1, 1).value = self.sheet_name
                            bk.create_sheet(self.sheet_name)

                        for i in range(1,sheet0.max_row+1):
                            for j in range(1,sheet0.max_column+1):
                                bk[self.sheet_name].cell(i,j).value = sheet0.cell(i,j).value

                        bk.save(self.which_file)

                        QMessageBox.information(self, "QAQ", "保存成功!")

                    else :#未保存
                        QMessageBox.information(self, "warning", "当前正在编辑一个详细内容，点击该按钮无效，请先保存至缓存或者放弃编辑!")
                # except Exception as e:
                #     QMessageBox.information(self, "error", "edit_pane保存为历史编辑失败:%s" % e)
                else :
                    QMessageBox.information(self, "warning", "暂无数据，点击该按钮无效！")
        except Exception as e:
            print("editpane 231416",e)



    #这只是放弃下面这个的编辑
    def giveup_edit(self):
        try:
            if self.have_data_or_not == 1:
                if self.edit_what_flag == 1:
                    #这个放弃编辑操作就是
                    if self.save_or_not == 0:   #没保存
                        self.save_or_not = 1#类是保存的效果
                        #并且把左边设置未已
                        self.save_or_not_label.setText("当前状态-已保存")
                        #把设置为不可以编辑
                        self.textEdit.setReadOnly(True)
                        #打开
                        #重新到edit_filw sheet1中读取数据
                        #注意每次点击保存所有 sheet0和sheet1里面的的内容都会改变
                        #打开sheet1，把sheet1的详细信息写道sheet0中
                        #注意只是这一行的，
                        #应该是打开sheet0读取详细信息
                        wb = load_workbook("edit_file.xlsx")
                        sheet0 =wb["当前编辑"]
                        #
                        #获取行数
                        print("1234567")
                        print(self.edit_pane_row)
                        if sheet0.cell(self.edit_pane_row + 2,sheet0.max_column).value != None:
                            self.textEdit.setPlainText(str(sheet0.cell(self.edit_pane_row + 2,sheet0.max_column).value))
                        else:#无数据
                            self.textEdit.setPlainText("暂无详细信息!")
                    else :
                        #这是刚保存，没有点击其他（编辑，禁止编辑）按你牛
                        QMessageBox.information(self, "warning", "该内容已经被保存，无法放弃编辑操作！")
                else:
                    QMessageBox.information(self, "warning", "当前状态为编辑详表格，点击该按钮无效！")
            else:
                QMessageBox.information(self, "warning", "暂无数据，点击该按钮无效！")
        except Exception as e:
            print("editpane giveup_edit",e)


if __name__ == "__main__":
    import sys
    app = QApplication(sys.argv)

    _edit_pane = EditPane()
    _edit_pane.show()

    sys.exit(app.exec_())

