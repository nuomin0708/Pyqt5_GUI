from PyQt5.QtCore import pyqtSignal,QUrl
from PyQt5.QtWidgets import QWidget, QMessageBox, QApplication
from openpyxl import load_workbook
from PyQt5.QtGui import QDesktopServices


from concrete_ui import Ui_Form
class ConcretePane(QWidget,Ui_Form):
    goto_menu_con_siganl = pyqtSignal()
    goto_view_pane_signal = pyqtSignal()

    def __init__(self, parent=None, *args, **kwargs):
        super().__init__(parent, *args, **kwargs)
        self.setupUi(self)
        #设置为不可以编辑
        self.textEdit.setReadOnly(True)
        self.attention_textEdit.setReadOnly(True)
        self.col_()

    def col_(self):
        self.name_col = 1
        self.type_col = 2
        self.id_col = 3
        self.resume_col = 4
        self.pos_col = 5
        self.native_col = 6
        self.link1_col = 7
        self.link2_col = 8
        self.link1_tool_col = 9
        self.link2_tool_col = 10
        self.at_col = 11
        self.con_col = 12

    def goto_menu_con_or_view(self):
        if self.goto_menu_or_view_flag == 0 : #返回到主页i按
            self.goto_menu_con_siganl.emit()
        else :
            self.goto_view_pane_signal.emit()

    #munu -->   1
    #sercher     0

    #名字和编号 self.name_label
    # 简述 self.resume_label
    #摆放位置   self.position_label
    #链接1 link1_btn
    #链接2 link2_btn
    #详细框 self.textEdit
    # def about_dailog_to_this_pane(self):
        #通过对话框到这个界面的
        #注意不满组的话不可以跳到这这个界面
        #可以 先不设置为空

    def some_canshu(self,str1,str2,int1,int2,goto_flag):
        self.file_name = str1
        self.sheet_name = str2
        self.con_row = int1
        self.menu_or_search = int2
        self.goto_menu_or_view_flag = goto_flag

    def write_data_to_textline(self): #这里的sheet是文件
        try:
            #先把之前的内容给清除
            self.name_label.clear()
            self.resume_label.clear()
            self.type_label.clear()
            self.position_label.clear()
            self.textEdit.clear()
            self.attention_textEdit.clear()
            self.native_label.clear()

            self.open_book = load_workbook(self.file_name)  # 打开table工作本
            self.sheetx = self.open_book[self.sheet_name]
            #名称设置为0行的
            #有一种是什么数据都没有的
            #直接把它设置为不可以点击
            #

            name = str(self.sheetx.cell(int(self.con_row) + int(self.menu_or_search), self.name_col).value)
            id = str(self.sheetx.cell(int(self.con_row) + int(self.menu_or_search), self.id_col).value)
            self.name_label.setText(name+"-"+id)
            # self.name_label.setText(str(sheetx.cell(int(con_row) + int(menu_or_search),1).value))
            # 简述
            if  self.sheetx.cell(int(self.con_row) + int(self.menu_or_search), self.resume_col).value != None:
                self.resume_label.setText("——"+str(self.sheetx.cell(int(self.con_row) + int(self.menu_or_search), self.resume_col).value))
            else :
                self.resume_label.setText("——"+"暂无简述")
            #摆放位置
            #型号
            if  self.sheetx.cell(int(self.con_row) + int(self.menu_or_search),self.type_col).value != None:
                self.type_label.setText(str(self.sheetx.cell(int(self.con_row) + int(self.menu_or_search),self.type_col).value))
            else:
                self.type_label.setText("暂无型号")
            if   self.sheetx.cell(int(self.con_row) + int(self.menu_or_search),self.pos_col).value != None:
                self.position_label.setText(str(self.sheetx.cell(int(self.con_row) + int(self.menu_or_search),self.pos_col).value))
            else :
                self.position_label.setText("暂无")
            if self.sheetx.cell(int(self.con_row) + int(self.menu_or_search),self.native_col).value != None:
                self.native_label.setText(str(self.sheetx.cell(int(self.con_row) + int(self.menu_or_search),self.native_col).value))
            else:
                self.native_label.setText("未知")
            # #链接1
            # self.link1_btn.setText(sheetx.cell(int(self.con_row) + 1, 3).value)
            # #链接2
            # self.link2_btn.setText(sheetx.cell(int(self.con_row) + 1, 4).value)
            #详细信息
            if  self.sheetx.cell(int(self.con_row) + int(self.menu_or_search), self.con_col).value != None:
                self.textEdit.setText(str(self.sheetx.cell(int(self.con_row) + int(self.menu_or_search), self.con_col).value))
            else:
                self.textEdit.setText("暂无详细信息！")
            #注意事项
            if  self.sheetx.cell(int(self.con_row) + int(self.menu_or_search),self. at_col).value != None:
                self.attention_textEdit.setText(str(self.sheetx.cell(int(self.con_row) + int(self.menu_or_search),self. at_col).value))
            else:
                self.attention_textEdit.setText("无")

        except Exception as e:
            QMessageBox.information(self, "error", "ConcretePane错误 %s" % e)
            print("con write_data_to_textline",e)

    def about_link(self):
        print("ok")
        if self.sender().text() == "链接1" :
            which_link = self.sheetx.cell(int(self.con_row) + int(self.menu_or_search),self.link1_col).value
            link_or_not = self.sheetx.cell(int(self.con_row) + int(self.menu_or_search),self.link1_tool_col).value
        else :
            which_link = self.sheetx.cell(int(self.con_row) + int(self.menu_or_search), self.link2_col).value
            link_or_not = self.sheetx.cell(int(self.con_row) + int(self.menu_or_search), self.link2_tool_col).value

        if link_or_not != None:
            if link_or_not == "是":  # 是 表示需要用到工具
                QMessageBox.information(self, "QAQ", "该网站访问速度可能会很慢！")
            reply = QMessageBox.question(self, "question", "确定去往该网址？", QMessageBox.Yes | QMessageBox.No)
            if reply == QMessageBox.Yes:
                # 去访问
               QDesktopServices.openUrl(QUrl(which_link))
        else:
            QMessageBox.information(self, "", "该链接暂不存在！")





if __name__ == "__main__":
    import sys
    app = QApplication(sys.argv)

    _concrete_pane = ConcretePane()
    _concrete_pane.show()

    sys.exit(app.exec_())

