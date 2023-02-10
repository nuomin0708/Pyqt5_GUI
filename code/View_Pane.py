from PyQt5.QtCore import Qt, pyqtSignal
from PyQt5.QtWidgets import QApplication, QWidget, QTableWidgetItem, QHeaderView, QAbstractItemView, QMessageBox
from openpyxl import load_workbook

from  view_ui import Ui_Form

class ViewPane(QWidget, Ui_Form):
    view_conc_signal = pyqtSignal(str,str,int,int,int)
    # file_name, sheet_name, excel_row, jiaozheng_row, goto_flag
    goback_to_menu_signal = pyqtSignal()
    edit_in_viewpane_signal = pyqtSignal(str,str,int)

    def __init__(self, parent=None, *args, **kwargs):
        super().__init__(parent, *args, **kwargs)
        self.setAttribute(Qt.WA_StyledBackground, True)
        self.setupUi(self)

        #自适应
        # 去掉序号列
        self.tableWidget.verticalHeader().setHidden(True)
        #
        # 表头高度
        self.tableWidget.horizontalHeader().setMinimumHeight(35)

        # 自适应
        self.tableWidget.horizontalHeader().setSectionResizeMode(QHeaderView.Stretch)
        self.tableWidget.horizontalHeader().resizeSections(QHeaderView.ResizeToContents)

        #
        # self.tableWidget.horizontalHeader().setSectionResizeMode(QHeaderView.Stretch)
        # self.tableWidget.verticalHeader().setSectionResizeMode(QHeaderView.Stretch)

        self.tableWidget.setSelectionBehavior(QAbstractItemView.SelectRows)
        self.tableWidget.setEditTriggers(QAbstractItemView.NoEditTriggers)  # 不可编辑

    def setName_ane_uitable_ViewPane(self,file_name,sheet_name):
        #上面的label   sheet_label
        self.file_name = file_name
        self.sheet_name = sheet_name
        if file_name  == "class_and_sheet.xlsx":
            XXX = "以下是正式文件中的%s" % sheet_name
        else :
            XXX = "以下是历史记录中的%s" % sheet_name
        self.sheet_label.setText(XXX)
        #把数据写入到表格中
        open_book = load_workbook(filename=file_name)  # 打开table工作本
        sheet = open_book[sheet_name]
        if sheet.cell(1, 1).value != None:  #
            self.have_data_or_not = 1  # 表示有数据
            x = 1  # 后面的详细信息也不显示
            row = sheet.max_row
            col = sheet.max_column
            lis = []
            self.tableWidget.setRowCount(row - 1)
            self.tableWidget.setColumnCount(col - x)
            for i in range(1, row + 1):
                for j in range(1, col + 1 - x):  # 列数   #第i行，第j列
                    if i == 1:
                        lis.append(str(sheet.cell(i, j).value))
                        if j == col + 1 - x - 1:
                            self.tableWidget.setHorizontalHeaderLabels(lis)
                    else:
                        # 有些excel单元格没有数据 就不用写入了
                        if sheet.cell(i, j).value != None:
                            item = QTableWidgetItem(str(sheet.cell(i, j).value))  # 获取单元格内容
                            # item = QTableWidgetItem(str(sheet.cell(i, j).value))  # 获取单元格内
                            self.tableWidget.setItem(i - 2, j - 1, item)
        else:
            self.have_data_or_not = 0  # 0表示煤油数据
            T_lis = [""]
            self.tableWidget.setHorizontalHeaderLabels(T_lis)
            self.tableWidget.setColumnCount(1)
            self.tableWidget.setRowCount(1)
            item = QTableWidgetItem("%s暂无数据" % sheet_name)
            self.tableWidget.setItem(0, 0, item)
            self.tableWidget.resizeColumnToContents(0)  # 自动列宽

    def view_conc(self):
        try:
            row = self.tableWidget.selectedItems()
            if len(row) > 0:  # 至少选中了一个数据
                wk = load_workbook(self.file_name)
                st = wk[self.sheet_name]
                if st.cell(1, 1).value != None:
                    row_enable = self.tableWidget.selectedItems()[0].row()
                    #去到具体
                    self.view_conc_signal.emit(self.file_name,self.sheet_name,row_enable, 2,1)
                    # self.emit_goto_concrete_signal.emit(, row_enable, 2)  # 第一个是sheet，第二个是行
                    #
                    # 加个两行
                else:
                    # 表头是固定的
                    QMessageBox.information(self, "warning", "该类暂无添加数据！")
            else:
                QMessageBox.information(self, "error", "请选中一个内容!")
        except Exception as e:
            QMessageBox.information(self, "error", "请选中一个内容! %s " % e)


    def goback_to_menu(self):
        self.goback_to_menu_signal.emit()

    def edit_cao(self):
        self.edit_in_viewpane_signal.emit( self.file_name ,self.sheet_name,1)
        #0 用于editpane1的返回按钮


if __name__ == '__main__':
    import sys
    app = QApplication(sys.argv)
    _view_pane = ViewPane()
    _view_pane.show()
    sys.exit(app.exec_())
