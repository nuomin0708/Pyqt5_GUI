
from PyQt5.QtCore import pyqtSignal, Qt
from PyQt5.QtWidgets import QWidget, QAbstractItemView, QMessageBox, QFileDialog, QTableWidgetItem,  \
    QApplication,QHeaderView
from openpyxl import Workbook,load_workbook
from menu_ui import Ui_Form
from Initialization import Initi

class MenuPane(QWidget, Ui_Form, Initi):

    select_new_or_old_class_signal = pyqtSignal()
    emit_path_to_main_signal = pyqtSignal(list,int)  #把外部文件sheet发出去
    emit_select_sheet_signal = pyqtSignal()
    open_pc_excel_file_signal = pyqtSignal(list)
    menupane_to_personalpane_signal = pyqtSignal()

    open_file_for_sheet0_signal = pyqtSignal(list,str)

    emit_goto_concrete_signal = pyqtSignal(str,str,int,int,int)  #第一个是文件名第一个为sheet名，第二个是当前row，第三个是矫正row
    #最后普一个in图标篇是从大‘
    goto_history_dialog_signal = pyqtSignal(str)

    old_lis_signal = pyqtSignal()
    add_new_signal = pyqtSignal()
    view_history_signal = pyqtSignal()
    one_signal = pyqtSignal(list,int)
    one_file_signal = pyqtSignal()

    old_or_history_signal = pyqtSignal(str)  #

    hide_input_dialog_signal = pyqtSignal()

    open_edit_pane_signal = pyqtSignal(str,str)

    open_lis_kuang_dialog_signal = pyqtSignal()

    new_dialog_again_signal = pyqtSignal()

    search_signal = pyqtSignal()
    k = 0
    lis = []
    lis_h = []
    def __init__(self, parent=None, *args, **kwargs):
        super().__init__(parent, *args, **kwargs)
        self.setAttribute(Qt.WA_StyledBackground, True)
        self.setupUi(self)
        # self.setWindowTitle = "信息查询"
        #设置左右两边按钮为扁平
        # self.left_btn.setFlat(True)
        # self.right_btn.setFlat(True)

        #把光标取消
        # self.lineEdit.setFocusPolicy(Qt.StrongFocus)
        # self.search_lineEdit.setFocus(True)
        # self.search_lineEdit.clearFocus()
        self.search_lineEdit.returnPressed.connect(self.search_cao)

        #去掉序号列
        self.tableWidget.verticalHeader().setHidden(True)
        #
        #表头高度
        self.tableWidget.horizontalHeader().setMinimumHeight(35)

        #自适应
        self.tableWidget.horizontalHeader().setSectionResizeMode(QHeaderView.Stretch)
        self.tableWidget.horizontalHeader().resizeSections(QHeaderView.ResizeToContents)

        #
        # self.tableWidget.horizontalHeader().setSectionResizeMode(QHeaderView.Stretch)
        # self.tableWidget.verticalHeader().setSectionResizeMode(QHeaderView.Stretch)

        self.tableWidget.setSelectionBehavior(QAbstractItemView.SelectRows)
        self.tableWidget.setEditTriggers(QAbstractItemView.NoEditTriggers)  #不可编辑

        self.tableWidget.horizontalHeader().setStyleSheet(
            "QHeaderView::section{background-color: rgb(220, 220, 220);font:12pt '宋体';color: black;};")

        self._app_data.setValue("title_color", None)

         # 第一次使用程序  建立class_and_sheet0 文件
        if self._app_data.value("first_use_GUI") == None:
            try:
                self._app_data.setValue("first_use_GUI", 1)
                self.build_book_and_sheet0()
                self.build_edit_book_and_sheet0()
            except Exception as e:
                QMessageBox.information(self, "error", "MenuPane初始化错误 %s" % e)
        else:
            try:
                self.read_sheet0_to_lis()
                self.init_btn()
            except Exception as e:
                QMessageBox.information(self, "error", "MenuPane初始化2错误 %s" % e)


    def closeEvent(self, event):
        reply = QMessageBox.question(self, "question", "确定退出系统？", QMessageBox.Yes | QMessageBox.No)
        if reply == QMessageBox.Yes:
            event.accept()
        else:
            event.ignore()



    def search_cao(self):
        try:
            if self.search_lineEdit == None:
                QMessageBox.information(self, "warning", "编号不能为空！")
            elif len(self.search_lineEdit.text().strip()) == 0:
                QMessageBox.information(self, "warning", "编号不能为空！")
            elif len(self.search_lineEdit.text().strip() ) == 5 :
                #去寻找
                print("ok")
                self.search_fun()
                # self.search_signal.emit()
                #去调用搜索函数
            else:
                QMessageBox.information(self, "warning", "输入不规范")
        except Exception as e:
            print("mennu search_cao",e)

    def search_fun(self):
        self.have_sheet_flag = 0
        self.have_data_flag =0
        book = load_workbook("class_and_sheet.xlsx")
        sheet0 = book["分类及编号"]
        #获取前两位
        # print(self.search_lineEdit.text().strip())
        # print(self.search_lineEdit.text().strip()[0:2])
        # if self.search_lineEdit()[]
        x = 3  #表示编号所在行数
        for p in range(2,sheet0.max_row + 1):
            if self.search_lineEdit.text().strip()[0:2] == str(sheet0.cell(p,2).value):
                #找到了这个sheet
                self.have_sheet_flag =1
                sheetx = book[str(sheet0.cell(p, 1).value)]

                #把行数发出去
                for i in range(2,sheetx.max_row + 1):
                    if self.search_lineEdit.text().strip() == str(sheetx.cell(i,x).value) :
                        self.have_data_flag=1
                        QMessageBox.information(self, "QAQ", "找到了！")
                        self.emit_goto_concrete_signal.emit("class_and_sheet.xlsx",str(sheet0.cell(p,1).value),i,0,0)  # 注意这里的行是excel里面的行
                        #
                        break

                if  self.have_data_flag == 0: #没有这个数据
                    QMessageBox.information(self, "warning", "该编号不存在！")

                break

        if self.have_sheet_flag == 0: #没有找到数据
            QMessageBox.information(self, "warning", "该编号不存在！")

    def build_book_and_sheet0(self):
        try:
            wb = Workbook()   #wb是工作本
            sheet = wb.active  #
            sheet['A1'] = '所有分类'
            sheet["B1"] = "sheet编号"
            sheet.title = '分类及编号'
            wb.save("./class_and_sheet.xlsx")  # 保存到硬盘
        except Exception as e:
            QMessageBox.information(self, "error", "MenuPane错误 %s" % e)

    def build_edit_book_and_sheet0(self):
        try:
            wb = Workbook()  # wb是工作本
            sheet = wb.active  #
            sheet['A1'] = '所有编辑记录'
            sheet.title = '编辑记录'
            wb.save("./edit_class_and_sheet.xlsx")  # 保存到硬盘
        except Exception as e:
            QMessageBox.information(self, "error", "MenuPane错误建立edit_class_and_sheet文件失败 %s" % e)

    def read_sheet0_to_lis(self):
        try:
            wb = load_workbook("class_and_sheet.xlsx")  # 外部文件
            st = wb.worksheets[0]
            for i in range(2,st.max_row+1):
                self.lis.append(st.cell(i,1).value)

            wb2 = load_workbook("edit_class_and_sheet.xlsx")
            st2 = wb2.worksheets[0]
            for i in range(2,st2.max_row +1 ):
                self.lis_h.append(st2.cell(i,1).value)
        except Exception as e:
            QMessageBox.information(self, "error", "MenuPane错误读取sheet0错误 %s" % e)

    def init_btn(self):
        try:#等于0时不用导入数据
            if len(self.lis) >=1:
                if len(self.lis) == 1:
                    self.btn1.setEnabled(True)
                    self.btn2.setEnabled(False)
                    self.btn3.setEnabled(False)
                    self.btn4.setEnabled(False)
                    self.btn1.setText(self.lis[0])
                    self.btn2.setText("")
                    self.btn3.setText("")
                    self.btn4.setText("")
                    #左右两边按钮
                    self.left_btn.setEnabled(False)
                    self.right_btn.setEnabled(False)
                elif len(self.lis) == 2:
                    self.btn1.setEnabled(True)
                    self.btn2.setEnabled(True)
                    self.btn3.setEnabled(False)
                    self.btn4.setEnabled(False)
                    self.btn1.setText(self.lis[0])
                    self.btn2.setText(self.lis[1])
                    self.btn3.setText("")
                    self.btn4.setText("")
                    self.left_btn.setEnabled(False)
                    self.right_btn.setEnabled(False)
                elif len(self.lis) == 3:
                    self.btn1.setEnabled(True)
                    self.btn2.setEnabled(True)
                    self.btn3.setEnabled(True)
                    self.btn4.setEnabled(False)
                    self.btn1.setText(self.lis[0])
                    self.btn2.setText(self.lis[1])
                    self.btn3.setText(self.lis[2])
                    self.btn4.setText("")
                    self.left_btn.setEnabled(False)
                    self.right_btn.setEnabled(False)
                else:  # 四个或四个以上
                    # 失能
                    self.btn1.setEnabled(True)
                    self.btn2.setEnabled(True)
                    self.btn3.setEnabled(True)
                    self.btn4.setEnabled(True)
                    self.enable_disable_left_right_btn(self.k)
                    self.display_four_btn_text(self.k)
            else:
                self.btn1.setEnabled(False)
                self.btn2.setEnabled(False)
                self.btn3.setEnabled(False)
                self.btn4.setEnabled(False)
                self.enable_disable_left_right_btn(self.k)
                self.btn1.setText("")
                self.btn2.setText("")
                self.btn3.setText("")
                self.btn4.setText("")
                self.left_btn.setEnabled(False)
                self.right_btn.setEnabled(False)
        except Exception as e:
            print("menu pane init_btn ",e)
            QMessageBox.information(self, "error", "MenuPane错误初始化四个按钮内容错了 %s" % e)


    def enable_disable_left_right_btn(self, k):
        try:
            if k == 0:
                self.right_btn.setEnabled(False)
            if k == len(self.lis) - 4:
                self.left_btn.setEnabled(False)
            if k != 0:
                self.right_btn.setEnabled(True)
            if k != len(self.lis) - 4:
                self.left_btn.setEnabled(True)
        except Exception as e:
            print("menu pane enable_disable_left_right_btn",e)
            QMessageBox.information(self, "error", "MenuPane错误左右使能失能部分错误 %s" % e)

    def display_four_btn_text(self, k):
        try:
            self.btn1.setText(self.lis[k])
            self.btn2.setText(self.lis[k + 1])
            self.btn3.setText(self.lis[k + 2])
            self.btn4.setText(self.lis[k + 3])
        except Exception as e:
            print("meuu pane display_four_btn_text",e)
            print("display_four_btn_text",e)
            print(self.lis[k])
            # print(self.lis[k+1])
            # print(self.lis[k+2])
            QMessageBox.information(self, "error", "MenuPane错误表上方类的显示，变化错误 %s" % e)

    def left_btn_cao(self):
        try:
            self.push_left_right_btn_reset_color()
            self.k = self.k + 1
            self.enable_disable_left_right_btn(self.k)
            self.display_four_btn_text(self.k)
        except Exception as e:
            QMessageBox.information(self, "error", "MenuPane错误左键槽部分错误 %s" % e)


    def push_left_right_btn_reset_color(self):
        # self.btn1.setStyleSheet('''QPushButton{font: 16pt "方正舒体";
        #                 background-color: blue;border-radius:15px;}''')
        # self.btn2.setStyleSheet('''QPushButton{font: 16pt "方正舒体";
        #                         background-color: blue;border-radius:15px;}''')
        # self.btn3.setStyleSheet('''QPushButton{font: 16pt "方正舒体";
        #                        background-color: blue;border-radius:15px;}''')
        # self.btn4.setStyleSheet('''QPushButton{font: 16pt "方正舒体";
        #                                background-color: blue;border-radius:15px;}''')
        self.btn1.setStyleSheet('''QPushButton{	font: 14pt "宋体";
color: rgb(0, 0, 0);
	background-color: rgb(220, 220, 220);
border:0px}''')
        self.btn2.setStyleSheet('''QPushButton{	font: 14pt "宋体";
color: rgb(0, 0, 0);
	background-color: rgb(220, 220, 220);
border:0px}''')
        self.btn3.setStyleSheet('''	font: 14pt "宋体";
color: rgb(0, 0, 0);
	background-color: rgb(220, 220, 220);
border:0px;}''')
        self.btn4.setStyleSheet('''	font: 14pt "宋体";
color: rgb(0, 0, 0);
	background-color: rgb(220, 220, 220);
border:0px}''')

    def right_btn_cao(self):
        try:
            self.push_left_right_btn_reset_color()
            self.k = self.k - 1
            self.enable_disable_left_right_btn(self.k)
            self.display_four_btn_text(self.k)
        except Exception as e:
            QMessageBox.information(self, "error", "MenuPane错误右键槽部分错误 %s" % e)


    def change_four_btn_color_cao(self):  # 这个是四个按钮对应的槽
        #把tablewidget点击的
        self.tableWidget.setCurrentItem(None)
        if self._app_data.value("title_color") == None:  # 中间按钮第一次按下，这个每次到主菜单界面都会变成None
            self._app_data.setValue("title_color", self.sender())
            # self.sender().setStyleSheet(''' QPushButton{font: 16pt "方正舒体";
            #                           background-color: red;border-radius:15px;}  ''')  # 新的设置为红色
            # self.sender().setStyleSheet(''' QPushButton{font: 16pt "方正舒体";
            #         background-color: rgb(255, 85, 0);border-radius:15px;}  ''')  # 新的设置为红色
            self.sender().setStyleSheet(''' font: 14pt "宋体";
	color: rgb(0, 0, 0);background-color: rgb(181, 181, 181);border:0px ''')  # 新的设置为红色
            self.form_jump_pack(self.sender().text())
            #把此时的文件找到
            self.pitch_file = self.sender().text()
            print(self.pitch_file)

        else:
            if self._app_data.value != self.sender():  # 不是同一个按钮
                former = self._app_data.value("title_color")  # 之前的
                former.setStyleSheet(''' QPushButton{ 	font: 14pt "宋体";
color: rgb(0, 0, 0);
	background-color: rgb(220, 220, 220);
border:0px}
                                ''')  # 之前的设置为蓝色
                # former.setStyleSheet(''' QPushButton{ font: 16pt "方正舒体";
                # background-color: blue;border-radius:15px;}
                # ''')  # 之前的设置为蓝色
                # former.setStyleSheet(''' QPushButton{ font: 16pt "方正舒体";
                #               background-color: rgb(234, 255, 94);border-radius:15px;}
                #               ''')  # 之前的设置为蓝色
                self._app_data.setValue("title_color", self.sender())  # 这次点击保存到ini文件中
                # self.sender().setStyleSheet(''' QPushButton{font: 16pt "方正舒体";
                #           background-color: red;border-radius:15px;}  ''')  # 新的设置为红色
                # self.sender().setStyleSheet(''' QPushButton{font: 16pt "方正舒体";
                #    background-color: rgb(255, 85, 0);border-radius:15px;}  ''')  # 新的设置为红色
                self.sender().setStyleSheet('''font: 14pt "宋体";
	color: rgb(0, 0, 0);background-color: rgb(181, 181, 181);border:0px   ''')  # 新的设置为红色
                self.form_jump_pack(self.sender().text())
                self.pitch_file = self.sender().text()
                print(self.pitch_file)

    def write_new_class_to_normal_file_sheet0(self, new_sheet_name):
        try:
            #先在正式文件中判断有没有重复，没有重复的话就写入
            self.register_repeat_flag = 0  # 新类未重复标志
            # wb = load_workbook("class_and_sheet.xlsx")
            # sheet0 = wb["分类及编号"]
            #
            if len(self.lis)>0:
                for i in range(len(self.lis)):
                    if new_sheet_name == self.lis[i]:
                        self.register_repeat_flag = 1  #在正式文件中有，再进一步查看有无编辑历史
                        break
                        #不一定在sheet0存在就存在编辑历史，要在编辑界面中点击了保存此次编辑才有
            if   self.register_repeat_flag == 1 :
                # wb2 = load_workbook("edit_class_and_sheet.xlsx")
                # st = wb2.worksheets[0]
                # self.lis_h = st.cell(st.max_row,1).value  # 获取所有的sheet名

                for i in range(len(self.lis_h)):
                    if new_sheet_name == self.lis_h[i]:
                        self.register_repeat_flag = 2

                if self.register_repeat_flag == 2 :
                        QMessageBox.information(self, "error", "%s已存在编辑历史，请选择'编辑历史'再次编辑!" % new_sheet_name)
                        #去到编辑历史
                        self.goto_history_dialog_signal.emit(new_sheet_name)

                if  self.register_repeat_flag ==1:  #没有编辑历史
                    #会自动跳出
                    #所以这里弄一个信号 再次调用 New——dialog
                    QMessageBox.information(self, "error", "此类已存在，请另取名字")
                    self.new_dialog_again_signal.emit()

            #没有该数据 写入到lass
            if self.register_repeat_flag == 0:  # 未在正式文件中方重复
                #新建到正式文件中
                workbook1 = load_workbook("class_and_sheet.xlsx")
                st1 = workbook1.worksheets[0]
                st1.cell(st1.max_row+1,1).value = str(new_sheet_name)
                workbook1.create_sheet(new_sheet_name)

                #把这个心机哪得文件改成全局变量 要用得
                self.new_n = new_sheet_name
                workbook1.save("class_and_sheet.xlsx")

                #立马跟新
                self.lis = []  # 先清空
                wb = load_workbook("class_and_sheet.xlsx")
                st =wb.worksheets[0]
                for i in range(2, st.max_row + 1):
                    self.lis.append(st.cell(i, 1).value)
                print("ag123", self.lis)
                # 按钮重新初始化
                self.init_btn()

                reply = QMessageBox.question(self, "question", "添加分类成功，是否立即添加数据？", QMessageBox.Yes | QMessageBox.No)

                if reply == QMessageBox.Yes:

                    self.file666, _ = QFileDialog.getOpenFileName(self, "选取文件", "./", "Excel Files (*.xlsx)")
                    if self.file666.strip() != "":
                        self.book666 = load_workbook(self.file666)
                        sheets = self.book666.get_sheet_names()  # 获取所有
                        self.emit_path_to_main_signal.emit(sheets,1)  #这个发射到editpane里
                        #1表示外企
                        #1表示当个sheet

        except Exception as e:
            QMessageBox.information(self, "error", "MenuPane错误2 %s" % e)

    def open_pc_excel_file(self):
        file, _ = QFileDialog.getOpenFileName(self, "选取文件", "./", "Excel Files (*.xlsx)")
        if file.strip() != "":
            book = load_workbook(file)
            sheets= book.get_sheet_names()  # 获取所有
            self.open_pc_excel_file_signal.emit(sheets)   #把sheet(列表）发出去

    def add_class_enable_left_right_btn(self):
        if len(self.lis) == 1:
            self.btn1.setText(self.lis[0])
        elif len(self.lis) == 2:
            self.btn1.setEnabled(True)
            self.btn2.setEnabled(True)
            self.btn1.setText(self.lis[0])
            self.btn2.setText(self.lis[1])
        elif len(self.lis) == 3:
            self.btn1.setEnabled(True)
            self.btn2.setEnabled(True)
            self.btn3.setEnabled(True)
            self.btn1.setText(self.lis[0])
            self.btn2.setText(self.lis[1])
            self.btn3.setText(self.lis[2])
        else:
            self.btn1.setEnabled(True)
            self.btn2.setEnabled(True)
            self.btn3.setEnabled(True)
            self.btn4.setEnabled(True)
            self.display_four_btn_text(self.k)
            self.enable_disable_left_right_btn(self.k)

    def form_jump_pack(self, position_text):
        try:
            # x = 1 #这个x是用来控制多少是不显示在menu界面的
            x = 6
            lis = []
            wb = load_workbook("class_and_sheet.xlsx")
            for i in range(1,wb.worksheets[0].max_row+1):
                if position_text == wb.worksheets[0].cell(i,1).value :
                    sheetx = wb[position_text]
                    #通过索引或者名字找到
                    if  sheetx.cell(1,1).value != None:  #说明有数据
                        self.tableWidget.setRowCount(sheetx.max_row - 1)
                        self.tableWidget.setColumnCount(sheetx.max_column-x)  #最后一个详细信息不展示
                        for i in range(1,sheetx.max_row+1):  # 把对应表格的数据拿出来
                            for j in range(1,sheetx.max_column+1-x):  # 列数   #第i行，第j列  #最后的详细信息不展示
                                if i == 1:
                                    lis.append(str(sheetx.cell(i, j).value))
                                    if j == sheetx.max_column+1-x-1:  # 行都有了
                                        self.tableWidget.setHorizontalHeaderLabels(lis)
                                else:
                                    if sheetx.cell(i, j).value != None:
                                        item = QTableWidgetItem(str(sheetx.cell(i, j).value))  # 获取单元格内容
                                        # item = QStandardItem(sheetx.cell(i, j).value)  # 获取单元格内容
                                        self.tableWidget.setItem(i - 2, j-1, item)
                    else:  # 空的
                        T_lis = [""]
                        self.tableWidget.setHorizontalHeaderLabels(T_lis)
                        self.tableWidget.setColumnCount(1)
                        self.tableWidget.setRowCount(1)
                        item = QTableWidgetItem("%s暂无12数据" % position_text)
                        self.tableWidget.setItem(0, 0, item)
                        self.tableWidget.resizeColumnToContents(0)  # 自动列宽
        except Exception as e:
            print("menu jump",e)
            # QMessageBox.information(self, "error", "MenuPane错误jump失败 %s" % e)

    def goto_concrete(self):
        try:
            row = self.tableWidget.selectedItems()
            if len(row)> 0:#至少选中了一个数据
                #
                #先看这个有没有数据
                wk = load_workbook("class_and_sheet.xlsx")
                st =wk[self.pitch_file]
                if st.cell(1,1).value != None:
                    row_enable = self.tableWidget.selectedItems()[0].row()
                    self.emit_goto_concrete_signal.emit("class_and_sheet.xlsx",self.pitch_file,row_enable,2,0)  #第一个是sheet，第二个是行
                    #
                    #加个两行
                else:
                    #表头是固定的
                    QMessageBox.information(self, "warning", "该类暂无添加数据！")
            else:
                QMessageBox.information(self, "error", "请选中一个内容!")
        except Exception as e:
            QMessageBox.information(self, "error", "请选中一个内容! %s " % e)

    #这是编辑历史 里的名称
    #修改名称
    def edit_name(self,file_class,sheet_name,new_name):
        #请输入新的名字
        try:
            workbook = load_workbook(filename=file_class)
            wk = workbook.worksheets[0]
            for i in range(2,wk.max_row+1):
                if sheet_name == wk.cell(i, 1).value:
                    wk.cell(i, 1).value = new_name
                    updateSheet = workbook[sheet_name]
                    updateSheet.title = new_name
                    break
            workbook.save(file_class)
            self.lis = []
            workbook1 = load_workbook(filename=file_class)
            wk1 = workbook1.worksheets[0]
            for i in range(2, wk1.max_row + 1):
                self.lis.append(wk1.cell(i, 1).value)
            print(self.lis)
            self.init_btn()
            #把关闭输入框信号发射出去
            self.hide_input_dialog_signal.emit()
            QMessageBox.information(self, "QAQ", "修改成功!")
            #问用户是否去修改数据
            reply = QMessageBox.question(self, "question", "是否去修改表格数据？", QMessageBox.Yes | QMessageBox.No)
            if reply == QMessageBox.Yes :
                #打开对应的编辑界面
                self.open_edit_pane_signal.emit(file_class,new_name)

            #更新数据
        except Exception as e:
            QMessageBox.information(self, "error", "Menu修改名称%s " % e)

    def delete_object(self,file_class,sheet_name):
        try:
            #其中的表格数据也会被删除
            #先询问是否删除
            if file_class == "class_and_xlsx" :
                self.the_class = "已有文件"
            else :
                self.the_class = "编辑记录"
            reply = QMessageBox.question(self, "question", "确定删除{0}里的{1}？".format(self.the_class,sheet_name), QMessageBox.Yes | QMessageBox.No)
            if reply == QMessageBox.Yes:
                workbook = load_workbook(filename=file_class)
                #
                sheet0 = workbook.worksheets[0]
                for i in range(1, sheet0.max_row + 1):
                    if sheet0.cell(i, 1).value == sheet_name:
                        # 把这行删除
                        sheet0.delete_rows(i)  # 表示删除表格的第一行
                #删除sheet
                worksheet = workbook[sheet_name]
                workbook.remove(worksheet)
                workbook.save(file_class)

                # 跟新
                self.lis = []
                workbook1 = load_workbook(filename=file_class)
                wk1 = workbook1.worksheets[0]
                for i in range(2, wk1.max_row + 1):
                    self.lis.append(wk1.cell(i, 1).value)
                print(self.lis)
                self.init_btn()

                QMessageBox.information(self, "QAQ", "成功删除!")
                reply = QMessageBox.question(self, "question", "是否去查看剩余器材分类？", QMessageBox.Yes | QMessageBox.No)
                if reply == QMessageBox.Yes:
                    # 打开对应的编辑界面
                    self.open_lis_kuang_dialog_signal.emit()

        except Exception as e:
            QMessageBox.information(self, "QAQ", "删除失败%s " % e)

    def open_file_for_sheet0(self,str):
        try:
            sheets = []
            workbook = load_workbook(str)
            wd = workbook.worksheets[0]
            for i in range(2,wd.max_row+1):
                sheets.append(wd.cell(i,1).value)
            self.open_file_for_sheet0_signal.emit(sheets,str)
        except Exception as e:
            QMessageBox.information(self, "QAQ", "hieroy%s" %  e)

    def old_lis_cao(self):
        #发送
        self.old_or_history_signal.emit("class_and_sheet.xlsx")

    def add_new_cao(self):
        self.add_new_signal.emit()

    def view_history_cao(self):
        self.old_or_history_signal.emit("edit_class_and_sheet.xlsx")

    def one_cao(self):
        try:
            reply = QMessageBox.question(self, "question", "确定删除所有数据并导入新数据？" ,QMessageBox.Yes | QMessageBox.No)
            if reply == QMessageBox.Yes:
                # self.one_lis = []

                #最好还留有一个sheet

                # print("已删除")
                # print(workbook.get_sheet_names())
                #
                self.file12, _ = QFileDialog.getOpenFileName(self, "选取文件", "./", "Excel Files (*.xlsx)")
                #先显示里面有哪些sheet
                print(self.file12)

                #把sheet名发出去
                if self.file12.strip() != "":
                    #清空之前的
                    self.workbook123 = load_workbook(self.file12)
                    sheets123 = self.workbook123.get_sheet_names()  #获取所有sheet
                    # print(sheets1)
                    #所有的sheet可以通过 sheets1.title 发出去
                    self.one_signal.emit(sheets123,0)  #0表示一键导入
                    # 第三位得1表示dialoh打入
        except Exception as e:
            print("mehu ",e)

    def write_just_a_sheet_data_to_normal(self,sheet_name):
        #sheet_name 是选中的那个
        st = self.book666[sheet_name]
        #
        #打开正式文件得哪个新建得
        n_wb = load_workbook("class_and_sheet.xlsx")
        n_st = n_wb[self.new_n]
        for i in range(1,st.max_row+1):
            for j in range(1,st.max_column+1):
                n_st.cell(i,j).value = st.cell(i,j).value

        #并且在sheet0中写入编号
        sheet0 = n_wb["分类及编号"]

        #有可能是空文件 但这没关系
        #行数没有这么多
        #列数也没有这么多 可能
        x = 3
        if st.max_row >= 2 and st.max_column >= x:
            #注意刚才
            sheet0.cell(sheet0.max_row ,2).value = str(st.cell(2,3).value)[0:2]

        n_wb.save("class_and_sheet.xlsx")
        print("menu  write_just_a_sheet_data_to_normal 写入欧克")

        #跟新
        self.lis =[]
        workbook1 = load_workbook(filename="class_and_sheet.xlsx")
        wk1 = workbook1.worksheets[0]
        for i in range(2, wk1.max_row + 1):
            self.lis.append(wk1.cell(i, 1).value)
        self.init_btn()


    def one_to_file(self):
        try:
            #按下确定检后
            # 清空原有数据
            workbook = load_workbook('class_and_sheet.xlsx')
            sheet0 = workbook.worksheets[0]
            sheet0.delete_cols(1)
            sheet0.delete_cols(1)
            sheet0.cell(1, 1).value = "所有分类"
            sheet0.cell(1, 2).value = "sheet编号"

            # 通过名字移除
            for i in range(len(self.lis)):
                sheet = workbook.get_sheet_by_name(self.lis[i])
                workbook.remove(sheet)
            workbook.save("class_and_sheet.xlsx")

            wb = load_workbook(self.file12)  #外部文件
            # 先把wb里的额sheet找出来
            wb_sheet = wb.get_sheet_names()
            wb2 = load_workbook("class_and_sheet.xlsx")
            shet = wb2.worksheets[0]
            row = shet.max_row
            for p in range(len(wb_sheet)):
                shet.cell(row+ 1+p, 1).value = wb_sheet[p]
                wb2.create_sheet(wb_sheet[p])
                #写入数据
                for i in range(1,wb.worksheets[p].max_row+1):
                    for j in range(1,wb.worksheets[p].max_column+1) :
                        wb2[wb_sheet[p]] .cell(i,j).value = wb.worksheets[p].cell(i,j).value
            wb2.save("class_and_sheet.xlsx")
            QMessageBox.information(self, "QAQ", "添加成功！")

            #写入到self.lis中
            self.lis=[] #先清空
            self.k = 0
            for i in range(2,shet.max_row+1):
                self.lis.append(shet.cell(i,1).value)
            print("ag123",self.lis)
            #按钮重新初始化
            self.init_btn()

            #把sheet里面的写到sheet0中
            w_wb = load_workbook("class_and_sheet.xlsx")
            w_st0 = w_wb["分类及编号"]
            #通过sheet0获取
            # for i in range(len(self.lis)):
            #     w_st0.cell(i+2,)

            for i in range(2,w_st0.max_row +1):
                #第二行第二列
                #第2行第3列
                x = 3
                w_st0.cell(i,2).value = str(w_wb[w_st0.cell(i, 1).value].cell(2, x).value)[0:2]
                # print(str(w_wb[w_st0.cell(i, 1).value].cell(2, 2).value))
                # print(str(w_wb[w_st0.cell(i, 1).value].cell(2, 2).value)[0:2])

            print("ok")

            w_wb.save("class_and_sheet.xlsx")



            #获取所有的sheet
            # w_sheet = w_wb.get_sheet_names()  #列表
            # print( w_sheet)
            # for i in range()

            #把原来有的文件清除
            # self.tableWidget.clear()
            # t_lis = [""]
            # self.tableWidget.setHorizontalHeaderLabels(t_lis)
            # self.
            self.tableWidget.setColumnCount(0)
            self.tableWidget.setRowCount(0)
        except Exception as e:
            print("menu ahgs",e)

    def reset_cao(self):

        try:
            reply = QMessageBox.question(self, "question", "重置会删除所有添加的文件，确认重置？", QMessageBox.Yes | QMessageBox.No)
            #把三个文件的内容都清空
            if reply == QMessageBox.Yes:
                normal_file = load_workbook("class_and_sheet.xlsx")
                history_file = load_workbook("edit_class_and_sheet.xlsx")
                edit_file = load_workbook("edit_file.xlsx")

                #normal_file
                normal_sheet0 = normal_file["分类及编号"]
                normal_sheet0.delete_cols(1)
                normal_sheet0.delete_cols(1)
                # 删除sheet
                all_normal_sheet = normal_file.get_sheet_names()
                for i in range(len(all_normal_sheet)):
                    if str(all_normal_sheet[i]) != "分类及编号" :
                        normal_file.remove(normal_file.get_sheet_by_name(str(all_normal_sheet[i])))
                normal_sheet0['A1'] = '所有分类'
                normal_sheet0["B1"] = "sheet编号"
                normal_file.save("class_and_sheet.xlsx")

                #history_file
                history_sheet0 = history_file["编辑记录"]
                history_sheet0.delete_cols(1)
                # 删除sheet
                all_history_sheet = history_file.get_sheet_names()
                for i in range(len(all_history_sheet)):
                    if str(all_history_sheet[i]) != '编辑记录':
                        history_file.remove(history_file.get_sheet_by_name(str(all_history_sheet[i])))
                history_sheet0['A1'] = '所有编辑记录'
                history_file.save("edit_class_and_sheet.xlsx")

                edit_sheet0 = edit_file["当前编辑"]
                edit_sheet1 = edit_file["对比用的"]
                edit_sheet0_row = edit_sheet0.max_row
                edit_sheet1_row = edit_sheet1.max_row
                for i in range(1, edit_sheet0_row + 1):
                    edit_sheet0.delete_rows(1)
                for i in range(1, edit_sheet1_row + 1):
                    edit_sheet1.delete_rows(1)
                edit_file.save("edit_file.xlsx")

                # 按下确定检后

                # 写入到self.lis中
                self.lis = []  # 先清空
                print(len(self.lis))
                self.k = 0
                self.init_btn()

                #变回原来的颜色
                self.btn1.setStyleSheet("""font: 14pt "宋体";color: rgb(0, 0, 0);background-color: rgb(220, 220, 220);border:0px
""")
                self.btn2.setStyleSheet("""font: 14pt "宋体";color: rgb(0, 0, 0);background-color: rgb(220, 220, 220);border:0px
                """)
                self.btn3.setStyleSheet("""font: 14pt "宋体";color: rgb(0, 0, 0);background-color: rgb(220, 220, 220);border:0px
                """)
                self.btn4.setStyleSheet("""font: 14pt "宋体";color: rgb(0, 0, 0);background-color: rgb(220, 220, 220);border:0px
                """)


                #提示重置成功！
                QMessageBox.information(self, "QAQ", "重置成功！")
                #
                # 清空内容
                self.tableWidget.setColumnCount(0)
                self.tableWidget.setRowCount(0)
                # self.tableWidget.clear()
                # t_lis = [""]
                # self.tableWidget.setHorizontalHeaderLabels(t_lis)
        except Exception as e:
            print("menu pane reset_cao",e)

if __name__ == '__main__':
    import sys
    app = QApplication(sys.argv)
    _menu_pane = MenuPane()
    _menu_pane.show()
    sys.exit(app.exec_())
